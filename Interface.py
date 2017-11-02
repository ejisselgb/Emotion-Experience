import wx
import cv2
import cv2.cv as cv
import wx.lib.platebtn as platebtn
from wx.lib.buttons import GenBitmapButton
import threading
from string import split
import time
import sys
import glob, os
from string import split
from time import sleep
from PIL import Image
from os.path import basename
import pyfaces
from pyfaces import pyfaces
from openpyxl import Workbook
import xlsxwriter
from openpyxl.chart.label import DataLabelList
import wx.grid as gridlib
from openpyxl.chart import (
    PieChart,
    ProjectedPieChart,
    Reference
)
from imutils.video import VideoStream


class MainWindow(wx.Panel):#clase de la interfaz principal
    def __init__(self, parent,capture):
        wx.Panel.__init__(self, parent)
        self.panelResultados = PanelResultados(self)
        mainSizer = wx.BoxSizer(wx.VERTICAL)
        self.result = None
        self.inicioSesion = None
        #parent.EVT_CLOSE(self, self.OnCloseWindow)
        font = wx.Font(18, wx.MODERN, wx.NORMAL, wx.BOLD)
        mainSizer.AddSpacer(20)
        #Label
        self.label1 = wx.StaticText(self,label="Emotion Experience")
        self.label1.SetFont(font)
        labelBoxSizer1 = wx.BoxSizer(wx.HORIZONTAL)
        labelBoxSizer1.Add(self.label1, 0, wx.LEFT, 200)
        mainSizer.Add(labelBoxSizer1,0)
        mainSizer.AddSpacer(10)
        # video
        self.videoWarper = wx.StaticBox(self, label="Video",size=(640,480))
        videoBoxSizer = wx.StaticBoxSizer(self.videoWarper, wx.VERTICAL)
        videoFrame = wx.Panel(self, -1,size=(640,440))
        self.cap = ShowCapture(videoFrame, capture)
        self.cap.estado = "Listo para comenzar!"
        videoBoxSizer.Add(videoFrame,0)
        mainSizer.Add(videoBoxSizer,0)
        self.panelPorcentajes()
        #mainSizer.Add(estadoBoxSizer,0)
        mainSizer.AddSpacer(20)
        #Label
        label = wx.StaticText(self,label="Establecer tiempo")
        labelwc = wx.StaticText(self,label="Dispositivo de captura")
        labelBoxSizer = wx.BoxSizer(wx.HORIZONTAL)
        labelBoxSizer.Add(label, 0, wx.LEFT,105)
        labelBoxSizer.Add(labelwc, 0, wx.LEFT,195)
        mainSizer.Add(labelBoxSizer,0)
        mainSizer.AddSpacer(5)
        #Textbox
        textBoxSizer = wx.BoxSizer(wx.HORIZONTAL)
        self.inputBox = wx.TextCtrl(self)
        label2 = wx.StaticText(self,label="Seg")
        combobox = wx.ComboBox(self, -1, choices=webcams, value='Webcam 0', style=wx.CB_DROPDOWN, size=[150,50])
        combobox.Bind(wx.EVT_COMBOBOX, self.webcamSelect)
        textBoxSizer.Add(self.inputBox, 0, wx.LEFT , 105)
        textBoxSizer.Add(label2, 0, wx.LEFT, 10)
        textBoxSizer.Add(combobox, 0, wx.LEFT , 150)
        textBoxSizer.SetDimension(0,0,640,1200)
        mainSizer.Add(textBoxSizer,0)
        #mainSizer.AddSpacer(2)
        # button sesion
        buttonBoxSizer = wx.BoxSizer(wx.HORIZONTAL)
        b_sesion = wx.Button(self, wx.ID_CLOSE, "Comenzar")
        b_sesion.Bind(wx.EVT_BUTTON, self.sesion)
        buttonBoxSizer.Add(b_sesion,0,  wx.LEFT , 115)
        # button resultados
        b_resultados = wx.Button(self, wx.ID_CLOSE, "Generar Informe")
        b_resultados.Bind(wx.EVT_BUTTON, self.generarInforme)
        buttonBoxSizer.Add(b_resultados,0, wx.LEFT, 210)
        mainSizer.Add(buttonBoxSizer,0)
        mainSizer.AddSpacer(10)
        # button cerrar
        buttonBoxSizer2 = wx.BoxSizer(wx.HORIZONTAL)
        m_close1 = wx.Button(self, wx.ID_CLOSE, "Cerrar")
        m_close1.Bind(wx.EVT_BUTTON, self.OnClose)
        buttonBoxSizer2.Add(m_close1,0, wx.LEFT, 270)
        mainSizer.Add(buttonBoxSizer2,0)

        parent.Centre()
        self.Show()
        self.SetSizerAndFit(mainSizer)
        
    def OnClose(self, event): #Funcion que controla el boton de cerrar
        dlg = wx.MessageDialog(self, 
            "Realmente desea salir de la aplicacion?",
            "Confirm Exit", wx.OK|wx.CANCEL|wx.ICON_QUESTION)
        self.result = dlg.ShowModal()
        if self.result == wx.ID_OK & cv2.waitKey(1):
            #dlg.Destroy()
            self.Close()
            frame.Close()
            self.Destroy()
        #if cv2.waitKey(1):
            sys.exit()
            
    def webcamSelect(self, event):
        indexWebcam = event.GetSelection()
        self.cap.capture = cv2.VideoCapture(int(indexWebcam))

    def panelPorcentajes(self):
        """"""
        self.framer = wx.Frame(None,-1,'Porcentaje de emociones',size=(420, 320))
        self.panelsito = PanelResultados(self.framer)

        
    def generarInforme(self, event):
        self.book = Workbook()
        self.panelsito.grid.SetCellValue(1,1,str(self.cap.porcentajeFeliz)+'%')
        self.panelsito.grid.SetCellValue(2,1,str(self.cap.porcentajeMolesto)+'%')
        self.panelsito.grid.SetCellValue(3,1,str(self.cap.porcentajeNeutro)+'%')
        self.framer.Show()
        self.sheet = self.book.active
        self.sheet.title = 'Emociones-Tiempo'
        self.sheet2 = self.book.create_sheet()
        self.sheet2.title = 'Porcentaje'
        self.sheet3 = self.book.create_sheet()
        self.sheet3.title = 'Informacion General'

        rows = []
        rowsT = []
        rowsE = []
        varA = 'A'
        varB = 'B'

        # Hoja de excel Emociones/Tiempo
        self.sheet['A1'] = 'Tiempo (SEG)'
        self.sheet['B1'] = 'Nombre Emociones'
        for i in range(len(self.cap.tiempos)):
            varA = varA + str(i + 2)
            rowsT = int(self.cap.tiempos[i])
            self.sheet[varA] = rowsT
            varA = 'A'

        for e in range(len(self.cap.emociones)):
            varB = varB + str(e + 2)
            rowsE = self.cap.emociones[e]
            self.sheet[varB] = rowsE
            varB = 'B'

        # Hoja de excel Porcentaje

        rowsR = []

        rowsR = [
            ('Emociones', 'Repeticiones', 'Porcentaje'),
            ('Alegria', int(self.cap.felicidad), int(self.cap.porcentajeFeliz)),
            ('Disgusto', int(self.cap.molestia), int(self.cap.porcentajeMolesto)),
            ('Neutro', int(self.cap.neutralidad), int(self.cap.porcentajeNeutro)),
        ]


        for row in rowsR:
            self.sheet2.append(row)

        rowI = []
        rowI = [
            ('Tasa de error', 'N. Imagenes', 'Distancia Promedio', 'Duracion de la sesion'),
            (self.cap.porcentajeError, self.cap.total, self.cap.distanciaPromedio, str(self.inputBox.GetValue()) + ' s'),
        ]
        for row in rowI:
            self.sheet3.append(row)

        data = Reference(self.sheet2, min_col=2, min_row=1, max_row=5, max_col=4)
        cats = Reference(self.sheet2, min_col=1, min_row=2, max_row=5)

        chart = PieChart()
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.dataLabels = DataLabelList() 
        chart.dataLabels.showPercent = True 
        chart.title = "Porcentaje de emociones reconocidas"

        self.sheet2.add_chart(chart, "D7")



        self.book.save("./Informes/InformeEmotionExperience"+"_"+time.strftime("%d-%m-%Y")+"_"+time.strftime("%H-%M-%S")+".xlsx")
        self.cap.estadoInterfaz.SetLabel("El informa ha sido generado!, esperando nuevas instrucciones")
        
    def sesion(self, event): #Funcion que se llama para iniciar una sesion de captura de emociones    
        self.cap.emociones = []
        self.cap.tiempos = []
        self.cap.distancias = []
        self.inicioSesion = False
        tiempo = None
        direccion = None
        tiempo = self.inputBox.GetValue()
        if(tiempo != None and tiempo.isdigit()): # valida si el valor indicado es numero
        # start the app
            self.dlg2 = wx.DirDialog(self, "Escoge la carpeta donde se almacenaran las caras:",
                           style=wx.DD_DEFAULT_STYLE
                           )
            if self.dlg2.ShowModal() == wx.ID_OK:
                direccion = self.dlg2.GetPath()
            self.dlg2.Destroy()
            if direccion != None:
                self.cap.direccion = direccion
                self.threads = []
                self.cap.tiempoLimite = int(tiempo)
                self.cap.inSesion = True
                t = threading.Timer(float(tiempo), self.cap.timersito)#Setea iniciar el timer segun el tiempo puesto por el usuario
                t.start()
                self.cap.conTime = 0
                self.cap.numberImage = 0
                self.x = threading.Thread(target=self.cap.rec)
                self.threads.append(self.x)
                self.x.start()
                self.inicioSesion = True

        else:
            print("no digito un numero");
            
        
class PanelResultados(wx.Panel):
    def __init__(self, parent):
        """Constructor"""
        wx.Panel.__init__(self, parent=parent)
        self.pfeliz = '0'
        self.pdisgusto = '0'
        self.pneutro = '0'
        self.grid = gridlib.Grid(self)
        self.grid.CreateGrid(125,122)
        self.grid.SetDefaultColSize(160)
        self.grid.SetDefaultRowSize(50)
        self.grid.SetDefaultCellAlignment(wx.ALIGN_CENTRE, wx.ALIGN_CENTRE)
        self.grid.SetCellValue(0,0,'Emocion')
        self.grid.SetCellValue(0,1,'Porcentaje')
        self.grid.SetCellValue(1,0,'Alegria')
        self.grid.SetCellValue(2,0,'Disgusto')
        self.grid.SetCellValue(3,0,'Neutro')
        self.grid.SetCellValue(1,1,self.pfeliz+'%')
        self.grid.SetCellValue(2,1,self.pdisgusto+'%')
        self.grid.SetCellValue(3,1,self.pneutro+'%')
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.grid, 0, wx.EXPAND)
        self.SetSizer(sizer)

class ShowCapture(wx.Panel):#clase en donde se setea parametros de video y la captura de caras
    def __init__(self, parent, capture, fps=24):
        wx.Panel.__init__(self, parent, wx.ID_ANY, (0,0), (640,480))
        self.emocion = None
        self.faces = None
        self.direccion = None
        self.imgFace = None
        self.time = 0
        self.conTime = 0
        self.numberImage = 0
        self.cambioEstado = False
        self.cx = None
        self.cy = None
        self.cw = None
        self.ch = None
        self.wf = None # final width
        self.x = None
        self.y = None
        self.w = None
        self.h = None
        self.emociones = [] #emociones que son detectadas por cada imagen extraida
        self.distancias = [] #las distancias registradas por cada comparacion, para sacar porcentaje de error
        self.tiempos = []#los tiempos en que se produce cada imagen
        self.direccionCascade = "./haarcascade_frontalface_default.xml"
        self.out = None

        self.capture = capture
        ret, frame = self.capture.read()

        height, width = frame.shape[:2]

        parent.SetSize((width, height))

        self.frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        self.bmp = wx.BitmapFromBuffer(width, height, self.frame)

        self.timer = wx.Timer(self)
        self.timer.Start(1000./fps)

        self.Bind(wx.EVT_PAINT, self.OnPaint)
        self.Bind(wx.EVT_TIMER, self.NextFrame)

        #Estado
        font1 = wx.Font(10, wx.MODERN, wx.NORMAL, wx.NORMAL)
        self.estadoInterfaz = wx.StaticText(self, wx.ID_ANY, label="Listo para comenzar!", style=wx.ALIGN_CENTER)
        self.estadoInterfaz.SetFont(font1)
        estadoBoxSizer = wx.BoxSizer(wx.HORIZONTAL)
        estadoBoxSizer.Add(self.estadoInterfaz,0,wx.LEFT,10)

        self.out = cv2.VideoWriter('./video/output.avi', -1, 8, (640,480))
        


    def OnPaint(self, evt):
        dc = wx.BufferedPaintDC(self)
        dc.DrawBitmap(self.bmp, 0, 0)


    def NextFrame(self, event):

        ret, frame = self.capture.read()

        if self.out != None:
            if ret==True:
                # write the flipped frame
                self.out.write(frame)
                #cv2.imshow('frame',frame)

            # Release everything if job is finished
            #self.capture.release()
            #self.out.release()
            #cv2.destroyAllWindows()

        if ret:
            frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            faceCascade = cv2.CascadeClassifier(self.direccionCascade)
            self.gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            self.faces = faceCascade.detectMultiScale(
                frame,
                scaleFactor=1.1,
                minNeighbors=5,
                minSize=(30, 30),
                flags=cv2.cv.CV_HAAR_SCALE_IMAGE
            )
            
            # Draw a rectangle around the faces
            for (x, y, w, h) in self.faces:
                cv2.rectangle(frame, (x, y), (x+w, y+h), (0, 0, 0), 1)
                cv2.putText(frame,self.emocion, (x,y), cv2.FONT_HERSHEY_SIMPLEX, 0.4, 255)
                self.cx = x
                self.cy = y
                self.cw = w
                self.ch = h
            self.bmp.CopyFromBuffer(frame)
            self.Refresh()           

    def rec(self): #extrae cara en segundo plano
        self.direccionImagenes = [] #almacena las direcciones de las caras extraidas y preprocesadas
        start = time.time()
        self.cambioEstado = True
        self.estadoInterfaz.SetLabel("Capturando....")
        #print self.out
        while self.inSesion:#mientras la sesion se encuentre verdadera el programa estara capturando y extrayendo caras
                if self.cx != None:#si el valor de x del rectangulo que se dibuja en la cara existe o es diferente a nulo significa que el sistema ha detectado un rostro
                    self.cropPhoto()#funcion que se encarga de extraer la cara, preprocesarla y almacenarla en disco
                    self.direccionImagenes.append(self.nameFile)#se agrega la direccion de la imagen de la cara a la lista de caras extraidas
                    self.tiempos.append(round(time.time()-start))
                    self.numberImage = self.numberImage + 1 #numberImage se encarga de que cada imagen de rostro tenga un nombre distinto al anterior     
                else:#si no existe ningun rostro detectado, no se realiza nada, se espera a que aparezca uno en camara  
                   self.cx = None
                #se limpian los valores que corresponden al rectangulo, con el fin de evitar errores
                self.cx = None
                self.cy = None
                self.cw = None
                self.ch = None
                self.conTime = self.conTime+1
                print self.tiempos
                print self.conTime
                sleep(1)#sleep se encarga de parar este loop por cuantos segundos se requiera, esto define cuantos rostros se extraen segun el tiempo que se asigne 
        if self.inSesion == False:#empieza la comparacion de las caras extraidas con la base de datos
                print("se acabo la captura, empieza el proceso eigenfaces")
                self.estadoInterfaz.SetLabel("Reconociendo emociones....")
                #se llama a pyfaces que contiene el algoritmo para la comparacion y definicion de emociones, uso: (imagenAComparar,carpetaFeliz,carpetaNeutro,carpetaMolesto,numeroEigenFaces,umbralDeAceptacion)
                directorioAlegria = os.path.abspath('./bd/Alegria')
                directorioDisgusto = os.path.abspath('./bd/Disgusto')
                directorioNeutro = os.path.abspath('./bd/Neutro')
                self.pyf=pyfaces.PyFaces(self.direccionImagenes,directorioAlegria,directorioNeutro,directorioDisgusto,int(4),float(0.3))
                self.emociones = self.pyf.arregloEmociones#se agrega a la lista de emociones la emocion que acaba de ser detectada
                self.distancias = self.pyf.arregloDistancias#se agrega a la lista de distancias la distancia que corresponde a la emocion detectada
                print(self.emociones)
                print(self.distancias)
                self.felicidad = self.emociones.count("Alegria")#cuenta cuantas veces se encuentra feliz en la lista
                self.neutralidad = self.emociones.count("Neutro")#cuenta cuantas veces se encuentra neutro en la lista
                self.molestia = self.emociones.count("Disgusto")#cuenta cuantas veces se encuentra molesto en la lista
                #se imprimen los resultados de conteo
                print "Resultados"
                print "Alegria se presento "+str(self.felicidad)+" veces"
                print "self.Neutralidad se presento "+str(self.neutralidad)+" veces"
                print "Disgusto se presento "+str(self.molestia)+" veces"
                #se sacan los porcentajes de cada emocion segun el numero de caras extraidas durante el proceso
                self.total = self.felicidad + self.neutralidad + self.molestia
                if self.total != 0:
                    self.porcentajeFeliz = self.felicidad*100/self.total
                    self.porcentajeNeutro = self.neutralidad*100/self.total
                    self.porcentajeMolesto = self.molestia*100/self.total
                else:
                    self.porcentajeFeliz = 0
                    self.porcentajeNeutro = 0
                    self.porcentajeMolesto = 0
                #se imprimen los porcentajes
                print "Porcentajes emociones"
                print "Felicidad "+str(self.porcentajeFeliz)+"%"
                print "Neutralidad "+str(self.porcentajeNeutro)+"%"
                print "Molestia "+str(self.porcentajeMolesto)+"%"
                print "Porcentaje de error"
                self.distanciaPromedio = sum(self.distancias)/len(self.distancias)
                self.porcentajeError = self.distanciaPromedio
                print self.porcentajeError,"%"
                self.estadoInterfaz.SetLabel("El informe esta listo!, presione Generar Informe")

                #se vacia la lista de emociones, para posteriores sesiones
                #self.emociones = []
    
    def cropPhoto(self): #recorta cara y almacena
        #se extrae la seccion de la imagen que corresponde a la cara detectada
        self.imgFace = self.gray[self.cy:self.cy+self.ch, self.cx:self.cx+self.cw]
        #se setea el size deseado
        size = 125, 125
        #se realiza el reescalado de la imagen
        resize = cv2.resize(self.imgFace, size, interpolation = cv2.INTER_AREA);
        #se setea la direccion y nombre de la imagen que contiene la cara preprocesada
        self.nameFile = self.direccion + "/emotional"  + str(self.numberImage) + ".png"
        #se almacena en disco la imagen de la cara resultante
        imgf = cv2.imwrite(self.nameFile, resize)
        
    #metodo que es llamado segun el tiempo suministrado por el usuario en la interfaz    
    def timersito(self):
        self.inSesion = False#cuando la sesion es falsa el programa deja de capturar y guardar las caras y empieza el proceso de comparacion con la bd


index = 0
capture = cv2.VideoCapture(index)
webcams = []
while True:
    ret, frame = capture.read(0) 

    if ret:
        webcams.append('Webcam '+str(index))
        index = index+1
        capture = cv2.VideoCapture(index)
    else:
        capture = cv2.VideoCapture(0)
        capture.set(cv.CV_CAP_PROP_FRAME_WIDTH, 640)
        capture.set(cv.CV_CAP_PROP_FRAME_HEIGHT, 480)
        break
print(webcams)

capture.set(cv.CV_CAP_PROP_FRAME_WIDTH, 640)
capture.set(cv.CV_CAP_PROP_FRAME_HEIGHT, 480)
app = wx.App(False)
frame = wx.Frame(None,-1,'Emotion Experience',size=(660, 730))
panel = MainWindow(frame,capture)
frame.Show()
app.MainLoop()
